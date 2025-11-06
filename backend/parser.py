from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Iterable, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries

from .config import get_settings
from .database import Base, SessionLocal, engine
from .models import Seat, SeatStatus

settings = get_settings()

STAGE_RANGE = "I35:AH38"


@dataclass
class SeatRecord:
    seat_id: str
    floor: int
    excel_row: int
    excel_col: int
    zone: str
    tier: Optional[str]
    price: int
    status: SeatStatus
    layout_row: Optional[int] = None
    layout_col: Optional[int] = None


FLOOR_1_RANGE = "C9:AN32"
FLOOR_1_SKIP_ROWS = {19}
FLOOR_1_SKIP_COLS = {column_index_from_string("L"), column_index_from_string("AE")}

FLOOR_1_MIN_COL, FLOOR_1_MIN_ROW, FLOOR_1_MAX_COL, FLOOR_1_MAX_ROW = range_boundaries(FLOOR_1_RANGE)

FLOOR_1_LAYOUT_ROWS = [
    row
    for row in range(FLOOR_1_MAX_ROW, FLOOR_1_MIN_ROW - 1, -1)
    if row not in FLOOR_1_SKIP_ROWS
]
FLOOR_1_LAYOUT_ROW_MAP = {row: index + 1 for index, row in enumerate(FLOOR_1_LAYOUT_ROWS)}

FLOOR_1_COLUMN_CENTER_LEFT = column_index_from_string("U")
FLOOR_1_COLUMN_CENTER_RIGHT = column_index_from_string("V")

FLOOR_2_CENTER = "C1:AN4"
FLOOR_2_LEFT = "A1:A16"
FLOOR_2_RIGHT = "AP1:AP16"
FLOOR_2_FORCE_CELLS = {"W3"}

LEFT_BOUNDARY = column_index_from_string("L")
RIGHT_BOUNDARY = column_index_from_string("AE")


def _has_border(cell) -> bool:
    border = getattr(cell, "border", None)
    if not border:
        return False
    return any(getattr(getattr(border, side), "style", None) for side in ("left", "right", "top", "bottom"))


def _cell_color(cell) -> Optional[str]:
    fill = getattr(cell, "fill", None)
    if not fill:
        return None
    start = getattr(fill, "start_color", None)
    if not start:
        return None
    if getattr(start, "type", "rgb") == "rgb" and getattr(start, "rgb", None):
        return start.rgb
    if getattr(start, "type", None) == "indexed" and getattr(start, "index", None) is not None:
        try:
            from openpyxl.styles.colors import COLOR_INDEX

            idx = int(start.index)
            return COLOR_INDEX.get(idx)
        except Exception:
            return None
    if getattr(start, "type", None) == "theme" and getattr(start, "theme", None) is not None:
        rgb = getattr(start, "rgb", None)
        return rgb
    return None


def _zone_for_column(col_index: int, floor: int) -> str:
    if floor == 1:
        if col_index < LEFT_BOUNDARY:
            return "F1-LEFT"
        if col_index > RIGHT_BOUNDARY:
            return "F1-RIGHT"
        return "F1-CENTER"
    if floor == 2:
        if col_index < LEFT_BOUNDARY:
            return "F2-LEFT"
        if col_index > RIGHT_BOUNDARY:
            return "F2-RIGHT"
        return "F2-CENTER"
    return f"F{floor}-ZONE"


def _iter_cells(ws, ref: str):
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            yield row, col, ws.cell(row=row, column=col)


def parse_floor_one(ws) -> List[SeatRecord]:
    seats: List[SeatRecord] = []
    for row, col, cell in _iter_cells(ws, FLOOR_1_RANGE):
        if row in FLOOR_1_SKIP_ROWS:
            continue
        if col in FLOOR_1_SKIP_COLS:
            continue
        if not _has_border(cell):
            continue
        tier = settings.tier_for_color(_cell_color(cell))
        status = SeatStatus.BLOCKED if tier is None else SeatStatus.AVAILABLE
        zone = _zone_for_column(col, floor=1)
        col_letter = get_column_letter(col)
        seat_id = f"1-{row}-{col_letter}"
        price = settings.price_for_tier(tier)
        layout_row = FLOOR_1_LAYOUT_ROW_MAP.get(row)
        if col == FLOOR_1_COLUMN_CENTER_LEFT:
            layout_col = 1
        elif col == FLOOR_1_COLUMN_CENTER_RIGHT:
            layout_col = 2
        elif col < FLOOR_1_COLUMN_CENTER_LEFT:
            layout_col = 1 + (FLOOR_1_COLUMN_CENTER_LEFT - col) * 2
        elif col > FLOOR_1_COLUMN_CENTER_RIGHT:
            layout_col = 2 + (col - FLOOR_1_COLUMN_CENTER_RIGHT) * 2
        else:
            layout_col = None

        seats.append(
            SeatRecord(
                seat_id=seat_id,
                floor=1,
                excel_row=row,
                excel_col=col,
                zone=zone,
                tier=tier,
                price=price,
                status=status,
                layout_row=layout_row,
                layout_col=layout_col,
            )
        )
    return seats


def parse_floor_two(ws) -> List[SeatRecord]:
    seats: List[SeatRecord] = []
    for row, col, cell in _iter_cells(ws, FLOOR_2_CENTER):
        col_letter = get_column_letter(col)
        coordinate = f"{col_letter}{row}"
        if coordinate not in FLOOR_2_FORCE_CELLS and not _has_border(cell):
            continue
        tier = settings.tier_for_color(_cell_color(cell))
        status = SeatStatus.BLOCKED if tier is None else SeatStatus.AVAILABLE
        zone = _zone_for_column(col, floor=2)
        seat_id = f"2-{row}-{col_letter}"
        price = settings.price_for_tier(tier)
        seats.append(
            SeatRecord(
                seat_id=seat_id,
                floor=2,
                excel_row=row,
                excel_col=col,
                zone=zone,
                tier=tier,
                price=price,
                status=status,
                layout_row=row,
                layout_col=col,
            )
        )
    for ref in (FLOOR_2_LEFT, FLOOR_2_RIGHT):
        for row, col, cell in _iter_cells(ws, ref):
            if not _has_border(cell):
                continue
            zone = _zone_for_column(col, floor=2)
            col_letter = get_column_letter(col)
            seat_id = f"2-{row}-{col_letter}"
            seats.append(
                SeatRecord(
                    seat_id=seat_id,
                    floor=2,
                    excel_row=row,
                    excel_col=col,
                    zone=zone,
                    tier=None,
                    price=0,
                    status=SeatStatus.BLOCKED,
                    layout_row=row,
                    layout_col=col,
                )
            )
    return seats


def parse_excel(path: Path) -> List[SeatRecord]:
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    return parse_floor_one(ws) + parse_floor_two(ws)


def persist_seats(seats: Iterable[SeatRecord]) -> None:
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    session = SessionLocal()
    try:
        for record in seats:
            seat = Seat(
                seat_id=record.seat_id,
                floor=record.floor,
                excel_row=record.excel_row,
                excel_col=record.excel_col,
                zone=record.zone,
                tier=record.tier,
                price=record.price,
                status=record.status,
                layout_row=record.layout_row,
                layout_col=record.layout_col,
            )
            session.add(seat)
        session.commit()
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


def write_json(seats: Iterable[SeatRecord], output_path: Path) -> None:
    payload = [
        {
            **asdict(record),
            "status": record.status.value,
        }
        for record in seats
    ]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    excel_path = Path("data/\u5f69\u8272\u5e73\u9762\u56fe\u6309\u821e\u53f0.xlsx")
    seats = parse_excel(excel_path)
    persist_seats(seats)
    write_json(seats, settings.seats_json_path)
    print(f"Parsed seats: {len(seats)}")


if __name__ == "__main__":
    main()
