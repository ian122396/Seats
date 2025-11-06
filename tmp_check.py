from openpyxl import load_workbook
from pathlib import Path
path=Path('data/彩色平面图按舞台.xlsx')
wb=load_workbook(path, data_only=False)
ws=wb.active
for coord in ['V3','W3','X3']:
    cell=ws[coord]
    border=[getattr(getattr(cell.border, side), 'style', None) for side in ('left','right','top','bottom')]
    fill=getattr(cell.fill,'start_color',None)
    rgb=getattr(fill,'rgb',None)
    print(coord, 'border', border, 'rgb', rgb, 'value', cell.value)
